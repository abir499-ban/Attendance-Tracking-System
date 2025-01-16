from sheet_dimensions import start_row, end_row, subject_col_map, id_column, absentee_threshold
import openpyxl
import mailer


def saveBook(book):
    book.save("D:\\Python\\attendance_tracker.py\\Book1.xlsx")
    print("Book Updated Successfully")


def handle(subject_code, roll_absentees):
    try:

        book = openpyxl.load_workbook("D:\\Python\\attendance_tracker.py\\Book1.xlsx")
        sheet = book["Sheet1"]
        #   r = sheet.max_row
        #   c = sheet.max_column
        for r in range(start_row, end_row):
            cell = sheet.cell(row=r, column=id_column)
            if cell.value in roll_absentees:
                active_cell = sheet.cell(row=r, column=subject_col_map[subject_code])

                if active_cell.value == absentee_threshold - 2:

                    print(f"warning mail to {sheet.cell(row = r, column=6).value} be sent/ only 1 leave left for subject {subject_code}") 
                    msg = f'Warning for you !! As you are left with only one abseent in subject : ${subject_code}'
                    mailer.SendMail(email=sheet.cell(row=r, column=6).value, msg=msg, warning=True)

                elif active_cell.value == absentee_threshold:
                    print("Exam withdrawal mail to be sent / Penalty imposed")
                    msg = f'You have reached the Maximum Absentee limit in subject : ${subject_code} and hence will detained from attending any exams in this particular subject'
                    mailer.SendMail(email=sheet.cell(row=r, column=6).value, msg=msg, warning=False)
                    
                active_cell.value += 1
                saveBook(book)

    except RuntimeError as e:
        print(e)

    pass
